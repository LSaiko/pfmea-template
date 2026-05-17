"""
pfmea_generator.py
==================
Generates a complete AIAG-VDA 2019 pFMEA Excel workbook for a Class III
CNC Machining – Ti-6Al-4V Femoral Stem manufacturing process.

Portfolio piece: Stryker Manufacturing Transfer Project Engineer application.
Regulatory context: 21 CFR 820 (FDA QSR), ISO 13485, AIAG-VDA FMEA 4th Ed. 2019.

Author : [Your Name]
Version: 1.1.0
"""

import sys
# Ensure Unicode print output works on Windows (cp1252) without PYTHONUTF8=1
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# PALETTE  (GMP / medical-device aesthetic)
# ---------------------------------------------------------------------------
NAVY        = "1F3864"   # header background
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"  # alternating row shade
RED_FILL    = "FF0000"  # H  priority
AMBER_FILL  = "FFC000"  # M  priority
GREEN_FILL  = "92D050"  # L  priority
GOLD        = "D4AF37"  # section title accent
TITLE_GRAY  = "D9D9D9"  # section-title row background

FOOTER_TEXT = "AIAG-VDA FMEA 2019  |  21 CFR 820  |  ISO 13485"
PROCESS_NAME = "CNC Machining – Ti-6Al-4V Femoral Stem"

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size)

def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def color_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, columns):
    """Style a header row with navy fill and white bold text."""
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = hdr_font()
        cell.fill = navy_fill()
        cell.alignment = center_align(wrap=True)
        cell.border = thin_border()

def apply_data_row(ws, row_num, values, alt=False):
    """Write a data row with optional alternating shade."""
    fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = body_font()
        cell.fill = fill
        cell.alignment = left_align()
        cell.border = thin_border()

def write_section_title(ws, row_num, title, max_col):
    """Merge cells across a row and write a section title."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=13, color=NAVY)
    cell.fill = color_fill(TITLE_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()

def write_footer(ws, row_num, max_col):
    """Merge a footer row at the bottom of a sheet."""
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=FOOTER_TEXT)
    cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    cell.alignment = center_align()
    cell.fill = color_fill("EFEFEF")

def autofit_columns(ws, min_width=12, max_width=50):
    """Approximate column auto-fit based on cell content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

# ---------------------------------------------------------------------------
# AIAG-VDA 2019  —  ACTION PRIORITY LOGIC  (replaces legacy RPN)
# ---------------------------------------------------------------------------

def get_action_priority(s, o, d):
    """
    Implements the AIAG-VDA 2019 Action Priority (AP) table.
    Returns 'H', 'M', or 'L' and a plain-language rationale string.

    Key principle: Detection gaps are weighted heavily for patient-safety
    devices because an undetected failure reaching the patient is catastrophic
    regardless of occurrence rate.
    """
    if s >= 9:
        if d >= 7:
            return "H", "S≥9 + D≥7: undetected critical failure – patient risk"
        elif 4 <= d <= 6:
            if o >= 4:
                return "H", "S≥9 + D=4-6 + O≥4: frequent critical w/ weak detection"
            else:
                return "M", "S≥9 + D=4-6 + O<4: rare critical but detection acceptable"
        else:  # d <= 3
            return "M", "S≥9 + D≤3: strong detection mitigates critical severity"
    elif s >= 7:
        if d >= 7:
            if o >= 4:
                return "H", "S=7-8 + D≥7 + O≥4: high-severity, frequent, poorly detected"
            else:
                return "M", "S=7-8 + D≥7 + O<4: high-severity but infrequent"
        elif 4 <= d <= 6:
            if o >= 4:
                return "M", "S=7-8 + D=4-6 + O≥4: moderate risk profile"
            else:
                return "L", "S=7-8 + D=4-6 + O<4: controlled risk"
        else:
            return "L", "S=7-8 + D≤3: detection adequate for severity level"
    else:  # s <= 6
        if d >= 7 and o >= 6:
            return "M", "S≤6 + high O & D: frequency/detectability concern"
        else:
            return "L", "Lower severity with adequate controls"


# ---------------------------------------------------------------------------
# SHEET 1 – SCOPE & PLANNING
# ---------------------------------------------------------------------------

def build_sheet1(wb):
    ws = wb.create_sheet("1. Scope & Planning")
    MAX_COL = 4
    ws.row_dimensions[1].height = 30

    write_section_title(ws, 1, "STEP 1 – Scope & Planning  |  " + PROCESS_NAME, MAX_COL)

    fields = [
        ("Process Name",          PROCESS_NAME),
        ("Scope Start",           "Raw Ti-6Al-4V Bar Stock Receipt"),
        ("Scope End",             "Final Dimensional Inspection / Release"),
        ("Device Classification", "Class III – PMA (21 CFR 814)"),
        ("Standard Applied",      "AIAG-VDA FMEA 4th Edition 2019"),
        ("Regulatory Framework",  "21 CFR 820 (QSR)  |  ISO 13485:2016"),
        ("Device Description",    "Cementless Ti-6Al-4V Femoral Stem – Implantable Hip Prosthesis"),
        ("Team Members",          "MFG Transfer PE, Quality Engineer, Process Engineer, Supplier Quality, R&D"),
        ("FMEA Facilitator",      "[Your Name] – Manufacturing Transfer Project Engineer"),
        ("FMEA Number",           "PFMEA-HIP-001"),
        ("Revision",              "Rev A"),
        ("Date",                  str(date.today())),
        ("Next Review Date",      "6 months post-transfer validation"),
        ("Status",                "Draft – Internal Review"),
    ]

    hdr_cols = ["Field", "Detail", "", ""]
    apply_header_row(ws, 2, hdr_cols)

    for i, (field, value) in enumerate(fields, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        # Label cell
        c1 = ws.cell(row=i, column=1, value=field)
        c1.font = body_font(bold=True)
        c1.fill = fill
        c1.alignment = left_align(wrap=False)
        c1.border = thin_border()
        # Merge value across remaining columns
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=MAX_COL)
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font = body_font()
        c2.fill = fill
        c2.alignment = left_align(wrap=False)
        c2.border = thin_border()

    write_footer(ws, len(fields) + 4, MAX_COL)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    return ws


# ---------------------------------------------------------------------------
# SHEET 2 – STRUCTURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet2(wb):
    ws = wb.create_sheet("2. Structure Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 2 – Structure Analysis  |  System Hierarchy Breakdown", MAX_COL)

    columns = ["System Level", "Element Name", "Function Contribution"]
    apply_header_row(ws, 2, columns)

    rows = [
        ("Level 1 – System",     "Manufacturing Cell",       "CNC Machining production environment"),
        ("Level 2 – Subsystem",  "CNC Turning Station",      "Primary material removal and geometry creation"),
        ("Level 3 – Component",  "Cutting Tool (Insert)",    "Executes material removal; controls Ra and geometry"),
        ("Level 3 – Component",  "Workholding Fixture",      "Maintains part datum and prevents movement"),
        ("Level 3 – Component",  "Coolant Delivery System",  "Controls chip evacuation and thermal stability"),
        ("Level 3 – Component",  "CNC Part Program",         "Defines toolpath, feeds, speeds, and tolerances"),
        ("Level 3 – Component",  "Operator / Setup Tech",    "Executes setup, monitors process, responds to alarms"),
        ("Level 4 – Interface",  "CMM Inspection Station",   "In-process and final dimensional verification"),
        ("Level 4 – Interface",  "Electronic Traveler (DHR)","Enforces operation sequence and records results"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws


# ---------------------------------------------------------------------------
# SHEET 3 – FUNCTION ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet3(wb):
    ws = wb.create_sheet("3. Function Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 3 – Function Analysis  |  Intended Functions & Requirements", MAX_COL)

    columns = ["Process Element", "Intended Function", "Requirement Satisfied"]
    apply_header_row(ws, 2, columns)

    rows = [
        ("Cutting Tool (Insert)",
         "Remove material to achieve OD ±0.010 mm and Ra ≤ 0.8 µm on taper surface",
         "Device drawing tolerance stack; osseointegration surface finish spec"),
        ("Workholding Fixture",
         "Maintain part concentricity <0.005 mm TIR during all turning operations",
         "Geometric dimensioning: perpendicularity and runout callouts"),
        ("Coolant Delivery System",
         "Deliver coolant at ≥50 PSI to maintain <60°C tool-chip interface temperature",
         "Ti-6Al-4V machining process spec; tool life and metallurgical integrity"),
        ("CNC Part Program",
         "Execute validated toolpath per ECO-approved revision; enforce feed/speed limits",
         "Validated manufacturing process (IQ/OQ/PQ); 21 CFR 820.70(i) software control"),
        ("Operator / Setup Tech",
         "Load correct program revision, set datum offsets, perform first-article inspection",
         "Work instruction WI-CNC-001; operator qualification records OQR-2024"),
        ("CMM Inspection Station",
         "Verify OD, ID, taper angle, and length within print tolerance post-machining",
         "Inspection plan IP-HIP-003; acceptance criteria per device drawing Rev G"),
        ("Electronic Traveler (DHR)",
         "Enforce operation sequence gate checks; capture lot/SN, operator, and tool data",
         "21 CFR 820.184 Device History Record; ISO 13485 §7.5.3 traceability"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws


# ---------------------------------------------------------------------------
# SHEET 4 – FAILURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet4(wb):
    ws = wb.create_sheet("4. Failure Analysis")
    MAX_COL = 7

    write_section_title(ws, 1, "STEP 4 – Failure Analysis  |  Failure Modes, Effects & Causes", MAX_COL)

    columns = [
        "Process Step", "Failure Mode",
        "Effect on Patient / Downstream Process",
        "Severity (1-10)", "Root Cause",
        "Current Prevention Control", "Current Detection Control"
    ]
    apply_header_row(ws, 2, columns)

    rows = [
        ("CNC Turning – OD",
         "OD out of tolerance (oversize)",
         "Implant-stem mismatch; potential intraoperative fracture or non-union – patient harm",
         8,
         "Tool wear beyond limit; incorrect offset entry; thermal expansion of Ti",
         "Insert change interval SOP; offset verification at setup",
         "Post-op CMM inspection; SPC chart OD trend"),
        ("CNC Turning – Surface",
         "Surface roughness Ra > 0.8 µm on taper",
         "Reduced osseointegration; long-term implant loosening – re-operation risk",
         9,
         "Worn or chipped insert; inadequate coolant flow; excessive feed rate",
         "Profilometer check per lot; insert change per interval",
         "Profilometer measurement per part; visual inspection"),
        ("Workholding – Fixture",
         "Fixture slip / part shift mid-cycle",
         "Geometric error cascade (runout, perpendicularity OOS); scrap or undetected escape",
         7,
         "Improper clamping torque; fixture wear; incorrect locating pin seating",
         "Torque spec on setup sheet; periodic fixture calibration",
         "First-article runout check; operator pre-run confirmation"),
        ("Coolant Delivery",
         "Coolant pressure drop / flow loss",
         "Thermal damage to Ti substrate; metallurgical alteration; implant scrapped or escaped",
         9,
         "Clogged filter; pump failure; hose leak; operator valve error",
         "Coolant PM schedule; filter change interval",
         "Visual observation only – no automated interlock (GAP)"),
        ("CNC Program – Revision",
         "Wrong program revision loaded (obsolete program)",
         "Incorrect geometry machined; toolpath collision risk; patient dimension non-conformance",
         8,
         "Paper traveler allows manual program selection; no enforced program lock",
         "Program naming convention; operator training",
         "Setup verification checklist; first-article measurement"),
        ("Inspection – CMM Step",
         "CMM inspection step skipped / bypassed",
         "Non-conforming part released to sterilization and distribution – direct patient risk",
         7,
         "Production schedule pressure; traveler signature forgery; inadequate gate control",
         "Work instruction WI-CMM-002; supervisor sign-off",
         "DHR review at final release – detection is end-of-line only (GAP)"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))
        # Highlight severity ≥ 9 in amber
        s_cell = ws.cell(row=i, column=4)
        if isinstance(row_data[3], int) and row_data[3] >= 9:
            s_cell.fill = color_fill(AMBER_FILL)
            s_cell.font = body_font(bold=True)

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws


# ---------------------------------------------------------------------------
# SHEET 5 – RISK RATING  (AIAG-VDA AP – NOT legacy RPN)
# ---------------------------------------------------------------------------

def build_sheet5(wb):
    ws = wb.create_sheet("5. Risk Rating")
    MAX_COL = 7

    # ---- AIAG-VDA Compliance Note Box (rows 1-6) -------------------------
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
        value="⚠  AIAG-VDA 2019 Action Priority (AP) — NOT Traditional RPN")
    t.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    t.fill = color_fill(NAVY)
    t.alignment = center_align()

    note = (
        "AIAG-VDA 2019 replaces Risk Priority Number (RPN = S×O×D) with Action Priority (AP: High / Medium / Low). "
        "Rationale for medical devices: a failure with Severity 9 that is NEVER detected (D=10) must be actioned "
        "immediately regardless of occurrence rate — RPN arithmetic can mask this by averaging scores. "
        "AP weights Detection gaps more heavily, reflecting the principle that an undetected failure reaching "
        "a patient is catastrophic. The AP lookup table (not multiplication) drives prioritisation."
    )
    ws.merge_cells("A2:G4")
    n = ws.cell(row=2, column=1, value=note)
    n.font = Font(name="Arial", size=10, italic=True)
    n.fill = color_fill("FFF2CC")   # light yellow
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 60

    ws.merge_cells("A5:G5")
    sub = ws.cell(row=5, column=1,
        value="AP Lookup: S9-10+D7-10→H | S9-10+D4-6+O≥4→H | S7-8+D7-10+O≥4→H | All others→M or L")
    sub.font = Font(name="Arial", size=9, bold=True, color=NAVY)
    sub.fill = color_fill("D9E1F2")
    sub.alignment = center_align()

    # Blank spacer
    ws.row_dimensions[6].height = 8

    # ---- Section Title --------------------------------------------------
    write_section_title(ws, 7, "STEP 5 – Risk Rating  |  AIAG-VDA Action Priority Table", MAX_COL)

    columns = [
        "Failure Mode",
        "S  (Severity\n1–10)",
        "O  (Occurrence\n1–10)",
        "D  (Detection\n1–10)",
        "Action\nPriority",
        "AP Rationale",
        "Legacy RPN\n(Reference Only)"
    ]
    apply_header_row(ws, 8, columns)
    ws.row_dimensions[8].height = 35

    # Pre-filled ratings aligned with Sheet 4 failure modes
    # Format: (failure_mode_label, S, O, D)
    rated_failures = [
        ("OD out of tolerance (oversize)",          8, 4, 3),
        ("Surface roughness Ra > 0.8 µm",           9, 3, 2),
        ("Fixture slip / part shift mid-cycle",     7, 3, 4),
        ("Coolant pressure drop / flow loss",       9, 3, 5),
        ("Wrong program revision loaded",           8, 2, 3),
        ("CMM inspection step skipped",             7, 4, 7),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (fm, s, o, d) in enumerate(rated_failures, start=9):
        ap, rationale = get_action_priority(s, o, d)
        rpn = s * o * d   # shown for reference / comparison only
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        row_vals = [fm, s, o, d, ap, rationale, rpn]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font = body_font(bold=(col_idx == 5))
            cell.border = thin_border()
            cell.alignment = center_align(wrap=True) if col_idx in (2,3,4,5,7) else left_align()

            if col_idx == 5:   # AP column — colour coded
                cell.fill = color_fill(AP_FILL[ap])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap == "H" else "000000")
            elif col_idx == 7:  # Legacy RPN – muted styling
                cell.fill = base_fill
                cell.font = Font(name="Arial", size=9, italic=True, color="808080")
            else:
                cell.fill = base_fill

    write_footer(ws, len(rated_failures) + 10, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["F"].width = 45
    return ws


# ---------------------------------------------------------------------------
# SHEET 6 – OPTIMIZATION / ACTIONS
# ---------------------------------------------------------------------------

def build_sheet6(wb):
    ws = wb.create_sheet("6. Optimization")
    MAX_COL = 9

    write_section_title(ws, 1, "STEP 6 – Optimization / Corrective Actions  |  Risk Reduction Plan", MAX_COL)

    columns = [
        "Initial AP", "Failure Mode",
        "Action Description", "Action Owner",
        "Target Date",
        "Revised\nS", "Revised\nO", "Revised\nD",
        "Revised AP"
    ]
    apply_header_row(ws, 2, columns)
    ws.row_dimensions[2].height = 35

    actions = [
        ("H",  "Coolant pressure drop / flow loss",
         "Install in-line coolant flow sensor with machine interlock: auto-feed hold + alarm if PSI < 45. "
         "Validate per IQ/OQ protocol IQ-COOL-001. Update PFMEA and Control Plan CP-HIP-001.",
         "Process Engineer", "2026-Q3",  9, 2, 2),

        ("H",  "Wrong program revision loaded",
         "Implement electronic traveler (eDHR) with barcode-locked program selection. "
         "Only approved ECO revision loads; prior revisions auto-archived. Validate per 21 CFR 820.70(i). "
         "Training completion rate 100% before go-live.",
         "MFG Transfer PE + IT", "2026-Q4",  8, 2, 1),

        ("M",  "Surface roughness Ra > 0.8 µm",
         "Reduce insert change interval from 150 to 100 parts; implement profilometer check "
         "every 50 parts (sample basis). Update PM schedule and WI-CNC-001 Rev B.",
         "Quality Engineer", "2026-Q3",  9, 2, 1),

        ("H",  "CMM inspection step skipped",
         "Add mandatory CMM gate-check in eDHR workflow: part cannot advance until CMM "
         "pass record is uploaded. CMM data auto-linked to DHR lot record. "
         "Supervisor override requires Level III QE approval.",
         "Quality + MFG PE", "2026-Q4",  7, 3, 2),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (init_ap, fm, action, owner, target, rs, ro, rd) in enumerate(actions, start=3):
        rev_ap, _ = get_action_priority(rs, ro, rd)
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        for col_idx, val in enumerate(
            [init_ap, fm, action, owner, target, rs, ro, rd, rev_ap], start=1
        ):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin_border()
            cell.font = body_font(bold=(col_idx in (1, 9)))

            if col_idx in (1, 9):   # AP cells
                ap_val = val
                cell.fill = color_fill(AP_FILL[ap_val])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap_val == "H" else "000000")
                cell.alignment = center_align()
            elif col_idx in (6, 7, 8):  # S/O/D revised
                cell.fill = base_fill
                cell.alignment = center_align()
            else:
                cell.fill = base_fill
                cell.alignment = left_align()

        ws.row_dimensions[i].height = 70   # tall rows for action descriptions

    write_footer(ws, len(actions) + 4, MAX_COL)
    autofit_columns(ws, min_width=10, max_width=60)
    ws.column_dimensions["C"].width = 60
    return ws


# ---------------------------------------------------------------------------
# SHEET 7 – RESULTS SUMMARY
# ---------------------------------------------------------------------------

def build_sheet7(wb):
    ws = wb.create_sheet("7. Results Summary")
    MAX_COL = 4

    write_section_title(ws, 1, "STEP 7 – Results & Documentation Summary", MAX_COL)

    # ---- Metrics Table --------------------------------------------------
    ws.merge_cells("A2:D2")
    mhdr = ws.cell(row=2, column=1, value="pFMEA Metrics Summary")
    mhdr.font = hdr_font(size=11)
    mhdr.fill = navy_fill()
    mhdr.alignment = center_align()

    metrics = [
        ("Total Failure Modes Analyzed",         6),
        ("Initial High (H) Action Priority",     2),
        ("High AP After Actions Implemented",    0),
        ("Medium (M) AP – Initial",              4),
        ("Medium (M) AP – After Actions",        2),
        ("Low (L) AP – After Actions",           4),
        ("Total Corrective Actions Assigned",    4),
        ("Actions with Owner Assigned",          4),
        ("Actions with Target Date",             4),
    ]

    for i, (label, value) in enumerate(metrics, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        # Label (merge 3 cols)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = body_font(bold=True)
        lc.fill = fill
        lc.alignment = left_align(wrap=False)
        lc.border = thin_border()
        # Value
        vc = ws.cell(row=i, column=4, value=value)
        vc.font = body_font(bold=True)
        vc.fill = fill
        vc.alignment = center_align()
        vc.border = thin_border()

    # ---- Linked Documents Table -----------------------------------------
    doc_start = len(metrics) + 4

    ws.merge_cells(start_row=doc_start, start_column=1,
                   end_row=doc_start, end_column=MAX_COL)
    dhdr = ws.cell(row=doc_start, column=1, value="Linked Quality System Documents")
    dhdr.font = hdr_font()
    dhdr.fill = navy_fill()
    dhdr.alignment = center_align()

    doc_cols = ["Document Type", "Document Number", "Title / Description", "Status"]
    apply_header_row(ws, doc_start + 1, doc_cols)

    docs = [
        ("Control Plan",       "CP-HIP-001 Rev A",  "CNC Femoral Stem – Machining Control Plan",         "In Review"),
        ("Work Instruction",   "WI-CNC-001 Rev B",  "CNC Turning Setup & Operation – Ti-6Al-4V Stem",    "Released"),
        ("Inspection Plan",    "IP-HIP-003 Rev A",  "CMM Dimensional Inspection – Femoral Stem",          "Released"),
        ("IQ/OQ Protocol",     "IQ-COOL-001",       "Coolant Interlock System – Installation Qualification","Pending"),
        ("OQ/PQ Protocol",     "PQ-CNC-STEM-001",   "Process Validation – CNC Machining Transfer",        "Planned"),
        ("Risk Management",    "RM-HIP-2024-01",    "Device Risk Management File (ISO 14971)",             "In Review"),
        ("DFMEA",              "DFMEA-HIP-001",     "Design FMEA – Cementless Femoral Stem",              "Released"),
        ("SOP",                "SOP-QE-022",        "pFMEA Creation and Maintenance Procedure",           "Released"),
    ]

    for i, doc_row in enumerate(docs, start=doc_start + 2):
        apply_data_row(ws, i, doc_row, alt=(i % 2 == 0))

    # ---- Archive Location -----------------------------------------------
    arch_row = doc_start + len(docs) + 3
    ws.merge_cells(start_row=arch_row, start_column=1, end_row=arch_row, end_column=MAX_COL)
    arc = ws.cell(row=arch_row, column=1,
        value="📁  Archive Location:  [PLM System] → Projects → HIP-STEM-TRANSFER-2024 → FMEA → PFMEA-HIP-001_RevA.xlsx")
    arc.font = Font(name="Arial", size=10, italic=True, color=NAVY)
    arc.fill = color_fill("EBF3FB")
    arc.alignment = left_align(wrap=False)
    arc.border = thin_border()

    write_footer(ws, arch_row + 2, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["C"].width = 48
    return ws


# ---------------------------------------------------------------------------
# COVER / WORKBOOK SETUP
# ---------------------------------------------------------------------------

def set_workbook_properties(wb):
    """Set metadata visible in Excel's file properties panel."""
    wb.properties.title   = "pFMEA – CNC Femoral Stem (AIAG-VDA 2019)"
    wb.properties.subject = "Process FMEA – Class III Medical Device"
    wb.properties.creator = "Manufacturing Transfer PE"
    wb.properties.company = "Stryker / [Your Organization]"
    wb.properties.keywords = "pFMEA, AIAG-VDA, ISO 13485, 21 CFR 820, Ti-6Al-4V"
    wb.properties.description = (
        "AIAG-VDA 2019 pFMEA for CNC Machining of Ti-6Al-4V Femoral Stem. "
        "Class III PMA device. Manufacturing Transfer project."
    )


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def generate_pfmea(output_path: str = "pfmea_aiag_vda_2019.xlsx"):
    """
    Entry point: builds all 7 worksheets and saves the workbook.

    Args:
        output_path: File path for the generated Excel workbook.
    """
    wb = Workbook()
    # Remove the default blank sheet created by openpyxl
    wb.remove(wb.active)

    print("Building AIAG-VDA 2019 pFMEA workbook …")

    build_sheet1(wb)
    print("  ✓  Sheet 1: Scope & Planning")

    build_sheet2(wb)
    print("  ✓  Sheet 2: Structure Analysis")

    build_sheet3(wb)
    print("  ✓  Sheet 3: Function Analysis")

    build_sheet4(wb)
    print("  ✓  Sheet 4: Failure Analysis")

    build_sheet5(wb)
    print("  ✓  Sheet 5: Risk Rating (AIAG-VDA AP)")

    build_sheet6(wb)
    print("  ✓  Sheet 6: Optimization / Actions")

    build_sheet7(wb)
    print("  ✓  Sheet 7: Results Summary")

    set_workbook_properties(wb)

    wb.save(output_path)
    print(f"\n✅  Workbook saved → {output_path}")
    print("    Open in Microsoft Excel or LibreOffice Calc.")
    print("    Recommended: review Sheet 5 compliance note before distribution.")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "pfmea_aiag_vda_2019.xlsx"
    generate_pfmea(out)
