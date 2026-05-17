"""
pfmea_generator_spinal.py
=========================
Generates a complete AIAG-VDA 2019 pFMEA Excel workbook for a Class II
Injection Molding process — PEEK Lumbar Interbody Fusion Cage.

Portfolio piece: Manufacturing Transfer / Quality Engineering showcase.
Regulatory context: 21 CFR 820 (FDA QSR), ISO 13485, AIAG-VDA FMEA 4th Ed. 2019.

Author : [Your Name]
Version: 1.0.0
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# PALETTE
# ---------------------------------------------------------------------------
NAVY       = "1F3864"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_FILL   = "FF0000"
AMBER_FILL = "FFC000"
GREEN_FILL = "92D050"
TITLE_GRAY = "D9D9D9"

FOOTER_TEXT  = "AIAG-VDA FMEA 2019  |  21 CFR 820  |  ISO 13485"
PROCESS_NAME = "Injection Molding – PEEK Lumbar Interbody Fusion Cage"

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size)

def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def color_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, columns):
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = hdr_font()
        cell.fill = navy_fill()
        cell.alignment = center_align(wrap=True)
        cell.border = thin_border()

def apply_data_row(ws, row_num, values, alt=False):
    fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = body_font()
        cell.fill = fill
        cell.alignment = left_align()
        cell.border = thin_border()

def write_section_title(ws, row_num, title, max_col):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=13, color=NAVY)
    cell.fill = color_fill(TITLE_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()

def write_footer(ws, row_num, max_col):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=FOOTER_TEXT)
    cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    cell.alignment = center_align()
    cell.fill = color_fill("EFEFEF")

def autofit_columns(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

# ---------------------------------------------------------------------------
# AIAG-VDA 2019 ACTION PRIORITY
# ---------------------------------------------------------------------------

def get_action_priority(s, o, d):
    if s >= 9:
        if d >= 7:
            return "H", "S>=9 + D>=7: undetected critical failure – patient risk"
        elif 4 <= d <= 6:
            if o >= 4:
                return "H", "S>=9 + D=4-6 + O>=4: frequent critical w/ weak detection"
            else:
                return "M", "S>=9 + D=4-6 + O<4: rare critical but detection acceptable"
        else:
            return "M", "S>=9 + D<=3: strong detection mitigates critical severity"
    elif s >= 7:
        if d >= 7:
            if o >= 4:
                return "H", "S=7-8 + D>=7 + O>=4: high-severity, frequent, poorly detected"
            else:
                return "M", "S=7-8 + D>=7 + O<4: high-severity but infrequent"
        elif 4 <= d <= 6:
            if o >= 4:
                return "M", "S=7-8 + D=4-6 + O>=4: moderate risk profile"
            else:
                return "L", "S=7-8 + D=4-6 + O<4: controlled risk"
        else:
            return "L", "S=7-8 + D<=3: detection adequate for severity level"
    else:
        if d >= 7 and o >= 6:
            return "M", "S<=6 + high O & D: frequency/detectability concern"
        else:
            return "L", "Lower severity with adequate controls"

# ---------------------------------------------------------------------------
# SHEET 1 – SCOPE & PLANNING
# ---------------------------------------------------------------------------

def build_sheet1(wb):
    ws = wb.create_sheet("1. Scope & Planning")
    MAX_COL = 4
    ws.row_dimensions[1].height = 30

    write_section_title(ws, 1, "STEP 1 – Scope & Planning  |  " + PROCESS_NAME, MAX_COL)

    fields = [
        ("Process Name",          PROCESS_NAME),
        ("Scope Start",           "Raw PEEK Pellet Receipt & Material Verification"),
        ("Scope End",             "Final Dimensional Inspection & Lot Release"),
        ("Device Classification", "Class II – 510(k) (21 CFR 888.3070)"),
        ("Standard Applied",      "AIAG-VDA FMEA 4th Edition 2019"),
        ("Regulatory Framework",  "21 CFR 820 (QSR)  |  ISO 13485:2016  |  ASTM F2026"),
        ("Device Description",    "PEEK Lumbar Interbody Fusion Cage – Spinal Implant"),
        ("Team Members",          "Mold Engineer, Quality Engineer, Process Engineer, Supplier Quality, R&D"),
        ("FMEA Facilitator",      "[Your Name] – Manufacturing Transfer Project Engineer"),
        ("FMEA Number",           "PFMEA-SPINE-001"),
        ("Revision",              "Rev A"),
        ("Date",                  str(date.today())),
        ("Next Review Date",      "6 months post-transfer validation"),
        ("Status",                "Draft – Internal Review"),
    ]

    apply_header_row(ws, 2, ["Field", "Detail", "", ""])

    for i, (field, value) in enumerate(fields, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        c1 = ws.cell(row=i, column=1, value=field)
        c1.font = body_font(bold=True)
        c1.fill = fill
        c1.alignment = left_align(wrap=False)
        c1.border = thin_border()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=MAX_COL)
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font = body_font()
        c2.fill = fill
        c2.alignment = left_align(wrap=False)
        c2.border = thin_border()

    write_footer(ws, len(fields) + 4, MAX_COL)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    return ws

# ---------------------------------------------------------------------------
# SHEET 2 – STRUCTURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet2(wb):
    ws = wb.create_sheet("2. Structure Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 2 – Structure Analysis  |  System Hierarchy Breakdown", MAX_COL)
    apply_header_row(ws, 2, ["System Level", "Element Name", "Function Contribution"])

    rows = [
        ("Level 1 – System",     "Injection Molding Cell",          "Thermoplastic implant manufacturing environment"),
        ("Level 2 – Subsystem",  "Injection Molding Press",         "Melts and injects PEEK at controlled temp/pressure"),
        ("Level 3 – Component",  "Mold Tool (Cavity/Core)",         "Defines part geometry, surface finish, and tolerances"),
        ("Level 3 – Component",  "Barrel / Screw Assembly",         "Plasticizes PEEK; controls melt temperature and shear"),
        ("Level 3 – Component",  "Hot Runner / Gate System",        "Delivers melt uniformly; controls gate vestige height"),
        ("Level 3 – Component",  "Material Handling & Drying",      "Ensures PEEK moisture <0.02% before processing"),
        ("Level 3 – Component",  "Process Controller (HMI)",        "Enforces validated parameter recipe (IQ/OQ/PQ)"),
        ("Level 4 – Interface",  "CMM / Vision Inspection Station", "Dimensional and visual verification post-mold"),
        ("Level 4 – Interface",  "Electronic Traveler (DHR)",       "Records lot, material CoA, operator, and process params"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 3 – FUNCTION ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet3(wb):
    ws = wb.create_sheet("3. Function Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 3 – Function Analysis  |  Intended Functions & Requirements", MAX_COL)
    apply_header_row(ws, 2, ["Process Element", "Intended Function", "Requirement Satisfied"])

    rows = [
        ("Mold Tool (Cavity/Core)",
         "Form cage geometry to OD/ID and footprint tolerances per drawing; achieve Ra <= 1.6 um on endplates",
         "Device drawing dimensional stack; osseointegration surface spec per ASTM F2026"),
        ("Barrel / Screw Assembly",
         "Maintain PEEK melt temperature 380-400 C; prevent thermal degradation of polymer",
         "PEEK processing spec; material property retention per ISO 10993-1 biocompatibility"),
        ("Hot Runner / Gate System",
         "Deliver bubble-free melt shot; gate vestige height <= 0.3 mm above parting line",
         "Device drawing gate vestige callout; visual inspection acceptance criteria"),
        ("Material Handling & Drying",
         "Dry PEEK pellets to moisture <= 0.02% at 150 C for >= 4 hours before molding",
         "PEEK processing guideline; prevention of hydrolytic degradation and splay defects"),
        ("Process Controller (HMI)",
         "Lock and enforce validated parameter recipe; flag deviation >5% from setpoint",
         "Validated manufacturing process (IQ/OQ/PQ); 21 CFR 820.70(i) software control"),
        ("CMM / Vision Inspection",
         "Verify footprint, height, endplate geometry, and gate vestige against print tolerance",
         "Inspection plan IP-SPINE-002; acceptance criteria per device drawing Rev D"),
        ("Electronic Traveler (DHR)",
         "Gate operation sequence; capture material lot, CoA link, operator, and process params",
         "21 CFR 820.184 Device History Record; ISO 13485 section 7.5.3 traceability"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 4 – FAILURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet4(wb):
    ws = wb.create_sheet("4. Failure Analysis")
    MAX_COL = 7

    write_section_title(ws, 1, "STEP 4 – Failure Analysis  |  Failure Modes, Effects & Causes", MAX_COL)

    columns = [
        "Process Step", "Failure Mode",
        "Effect on Patient / Downstream Process",
        "Severity (1-10)", "Root Cause",
        "Current Prevention Control", "Current Detection Control"
    ]
    apply_header_row(ws, 2, columns)

    rows = [
        ("Injection Molding – Parting Line",
         "Flash / fin exceeding 0.1 mm at parting line",
         "Sharp edge risk during implant handling; potential particulate in surgical field",
         6,
         "Insufficient clamp force; worn parting line; mold damage",
         "Clamp force setpoint in validated recipe; mold PM schedule",
         "100% visual inspection per WI-VIS-003; AQL sampling CMM"),
        ("Injection Molding – Geometry",
         "Outer footprint or height OOS (dimensional deviation)",
         "Implant-endplate mismatch; subsidence risk; non-union – patient harm",
         8,
         "Mold wear / thermal expansion; incorrect recipe offset; shot-to-shot variation",
         "Recipe lock in HMI; mold qualification per OQ; SPC on critical dims",
         "CMM inspection per sample plan SP-SPINE-01; first-article 100%"),
        ("Material Handling",
         "Wrong PEEK grade loaded (incorrect molecular weight / filler)",
         "Incorrect mechanical properties; implant fracture or fatigue failure in vivo",
         9,
         "Manual material selection; similar pellet appearance; label misread",
         "Material labeling SOP; operator training",
         "Visual label check only – no barcode/CoA lock (DETECTION GAP)"),
        ("Injection Molding – Fill",
         "Short shot / incomplete mold fill",
         "Structural porosity in load-bearing cage; fracture risk under spinal loading",
         7,
         "Insufficient injection pressure; blocked gate; material viscosity variation",
         "Injection pressure setpoint and transfer position in validated recipe",
         "Visual inspection – porosity may be subsurface and missed (GAP)"),
        ("Gate / Trim",
         "Gate vestige height > 0.3 mm above parting surface",
         "Mechanical interference with endplate; surgeon re-work; particulate generation",
         7,
         "Gate freeze inconsistency; dwell time variation; manual trim variability",
         "Dwell time in recipe; gate vestige spec on drawing",
         "First-article caliper check; periodic sampling only"),
        ("Injection Molding – Surface",
         "Endplate surface texture OOS (Ra > 1.6 um or peak pattern missing)",
         "Reduced bone in-growth; long-term instability and re-operation risk",
         8,
         "Worn or damaged mold texture; contamination on cavity surface; improper mold release",
         "Mold surface inspection at PM; release agent SOP",
         "Profilometer sample check per lot – not 100% (DETECTION GAP)"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))
        s_cell = ws.cell(row=i, column=4)
        if isinstance(row_data[3], int) and row_data[3] >= 8:
            s_cell.fill = color_fill(AMBER_FILL)
            s_cell.font = body_font(bold=True)

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 5 – RISK RATING
# ---------------------------------------------------------------------------

def build_sheet5(wb):
    ws = wb.create_sheet("5. Risk Rating")
    MAX_COL = 7

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
        value="  AIAG-VDA 2019 Action Priority (AP) -- NOT Traditional RPN")
    t.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    t.fill = color_fill(NAVY)
    t.alignment = center_align()

    note = (
        "AIAG-VDA 2019 replaces RPN (S x O x D) with Action Priority (AP: High / Medium / Low). "
        "For spinal implants, an undetected material substitution (wrong PEEK grade) reaching a patient "
        "represents a catastrophic fracture risk regardless of how rarely it occurs. "
        "AP weights Detection gaps heavily — a D=7 with any S>=9 is automatically High priority."
    )
    ws.merge_cells("A2:G4")
    n = ws.cell(row=2, column=1, value=note)
    n.font = Font(name="Arial", size=10, italic=True)
    n.fill = color_fill("FFF2CC")
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 60

    ws.merge_cells("A5:G5")
    sub = ws.cell(row=5, column=1,
        value="AP Lookup: S9-10+D7-10->H | S9-10+D4-6+O>=4->H | S7-8+D7-10+O>=4->H | All others->M or L")
    sub.font = Font(name="Arial", size=9, bold=True, color=NAVY)
    sub.fill = color_fill("D9E1F2")
    sub.alignment = center_align()
    ws.row_dimensions[6].height = 8

    write_section_title(ws, 7, "STEP 5 – Risk Rating  |  AIAG-VDA Action Priority Table", MAX_COL)

    columns = [
        "Failure Mode",
        "S  (Severity\n1-10)",
        "O  (Occurrence\n1-10)",
        "D  (Detection\n1-10)",
        "Action\nPriority",
        "AP Rationale",
        "Legacy RPN\n(Reference Only)"
    ]
    apply_header_row(ws, 8, columns)
    ws.row_dimensions[8].height = 35

    rated_failures = [
        ("Flash / fin at parting line",                6, 4, 3),
        ("Outer footprint or height OOS",              8, 4, 5),
        ("Wrong PEEK grade loaded",                    9, 3, 7),
        ("Short shot / incomplete fill",               7, 3, 2),
        ("Gate vestige height > 0.3 mm",               7, 4, 4),
        ("Endplate surface texture OOS",               8, 5, 7),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (fm, s, o, d) in enumerate(rated_failures, start=9):
        ap, rationale = get_action_priority(s, o, d)
        rpn = s * o * d
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        row_vals = [fm, s, o, d, ap, rationale, rpn]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font = body_font(bold=(col_idx == 5))
            cell.border = thin_border()
            cell.alignment = center_align(wrap=True) if col_idx in (2, 3, 4, 5, 7) else left_align()

            if col_idx == 5:
                cell.fill = color_fill(AP_FILL[ap])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap == "H" else "000000")
            elif col_idx == 7:
                cell.fill = base_fill
                cell.font = Font(name="Arial", size=9, italic=True, color="808080")
            else:
                cell.fill = base_fill

    write_footer(ws, len(rated_failures) + 10, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["F"].width = 45
    return ws

# ---------------------------------------------------------------------------
# SHEET 6 – OPTIMIZATION
# ---------------------------------------------------------------------------

def build_sheet6(wb):
    ws = wb.create_sheet("6. Optimization")
    MAX_COL = 9

    write_section_title(ws, 1, "STEP 6 – Optimization / Corrective Actions  |  Risk Reduction Plan", MAX_COL)

    columns = [
        "Initial AP", "Failure Mode",
        "Action Description", "Action Owner",
        "Target Date",
        "Revised\nS", "Revised\nO", "Revised\nD",
        "Revised AP"
    ]
    apply_header_row(ws, 2, columns)
    ws.row_dimensions[2].height = 35

    actions = [
        ("H",  "Wrong PEEK grade loaded",
         "Implement material barcode scan at press hopper: eDHR locks out start until scanned lot "
         "matches approved CoA on traveler. Certificate of Analysis auto-linked to DHR. "
         "Validate per 21 CFR 820.70(i); 100% operator training before go-live.",
         "MFG PE + Quality", "2026-Q3",  9, 2, 2),

        ("H",  "Endplate surface texture OOS",
         "Install inline profilometer at ejection station: 100% Ra measurement, SPC alert at Ra > 1.4 um. "
         "Mold surface inspection added to PM checklist at every 5,000 shots. "
         "Update Control Plan CP-SPINE-001 and WI-MOLD-002 Rev B.",
         "Process Engineer", "2026-Q3",  8, 3, 3),

        ("M",  "Outer footprint or height OOS",
         "Add 100% vision system check for critical footprint and height at ejection. "
         "Non-conforming parts automatically rejected to quarantine bin. "
         "SPC charts for cavity-specific dimensional trending.",
         "Quality Engineer", "2026-Q4",  8, 2, 2),

        ("M",  "Gate vestige height > 0.3 mm",
         "Add gate vestige height to mandatory first-article CMM program (currently omitted). "
         "Increase periodic sampling from 1/lot to 5/lot until Cpk >= 1.67 demonstrated over 30 lots.",
         "Quality + Mold Eng", "2026-Q4",  7, 3, 2),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (init_ap, fm, action, owner, target, rs, ro, rd) in enumerate(actions, start=3):
        rev_ap, _ = get_action_priority(rs, ro, rd)
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        for col_idx, val in enumerate(
            [init_ap, fm, action, owner, target, rs, ro, rd, rev_ap], start=1
        ):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin_border()
            cell.font = body_font(bold=(col_idx in (1, 9)))

            if col_idx in (1, 9):
                ap_val = val
                cell.fill = color_fill(AP_FILL[ap_val])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap_val == "H" else "000000")
                cell.alignment = center_align()
            elif col_idx in (6, 7, 8):
                cell.fill = base_fill
                cell.alignment = center_align()
            else:
                cell.fill = base_fill
                cell.alignment = left_align()

        ws.row_dimensions[i].height = 70

    write_footer(ws, len(actions) + 4, MAX_COL)
    autofit_columns(ws, min_width=10, max_width=60)
    ws.column_dimensions["C"].width = 60
    return ws

# ---------------------------------------------------------------------------
# SHEET 7 – RESULTS SUMMARY
# ---------------------------------------------------------------------------

def build_sheet7(wb):
    ws = wb.create_sheet("7. Results Summary")
    MAX_COL = 4

    write_section_title(ws, 1, "STEP 7 – Results & Documentation Summary", MAX_COL)

    ws.merge_cells("A2:D2")
    mhdr = ws.cell(row=2, column=1, value="pFMEA Metrics Summary")
    mhdr.font = hdr_font(size=11)
    mhdr.fill = navy_fill()
    mhdr.alignment = center_align()

    metrics = [
        ("Total Failure Modes Analyzed",         6),
        ("Initial High (H) Action Priority",     2),
        ("High AP After Actions Implemented",    0),
        ("Medium (M) AP – Initial",              2),
        ("Medium (M) AP – After Actions",        2),
        ("Low (L) AP – After Actions",           4),
        ("Total Corrective Actions Assigned",    4),
        ("Actions with Owner Assigned",          4),
        ("Actions with Target Date",             4),
    ]

    for i, (label, value) in enumerate(metrics, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = body_font(bold=True)
        lc.fill = fill
        lc.alignment = left_align(wrap=False)
        lc.border = thin_border()
        vc = ws.cell(row=i, column=4, value=value)
        vc.font = body_font(bold=True)
        vc.fill = fill
        vc.alignment = center_align()
        vc.border = thin_border()

    doc_start = len(metrics) + 4
    ws.merge_cells(start_row=doc_start, start_column=1,
                   end_row=doc_start, end_column=MAX_COL)
    dhdr = ws.cell(row=doc_start, column=1, value="Linked Quality System Documents")
    dhdr.font = hdr_font()
    dhdr.fill = navy_fill()
    dhdr.alignment = center_align()

    apply_header_row(ws, doc_start + 1,
                     ["Document Type", "Document Number", "Title / Description", "Status"])

    docs = [
        ("Control Plan",     "CP-SPINE-001 Rev A", "Injection Molding – PEEK Spinal Cage Control Plan",          "In Review"),
        ("Work Instruction", "WI-MOLD-002 Rev B",  "Injection Molding Setup & Operation – PEEK Cage",            "Released"),
        ("Inspection Plan",  "IP-SPINE-002 Rev A", "CMM & Vision Inspection – Lumbar Interbody Cage",            "Released"),
        ("IQ/OQ Protocol",   "IQ-MOLD-003",        "Injection Press Qualification – PEEK Processing Parameters", "Pending"),
        ("PQ Protocol",      "PQ-SPINE-CAGE-001",  "Process Validation – PEEK Cage Injection Molding Transfer",  "Planned"),
        ("Risk Management",  "RM-SPINE-2024-02",   "Device Risk Management File (ISO 14971) – Fusion Cage",      "In Review"),
        ("DFMEA",            "DFMEA-SPINE-001",    "Design FMEA – PEEK Lumbar Interbody Fusion Cage",            "Released"),
        ("SOP",              "SOP-QE-022",         "pFMEA Creation and Maintenance Procedure",                   "Released"),
    ]

    for i, doc_row in enumerate(docs, start=doc_start + 2):
        apply_data_row(ws, i, doc_row, alt=(i % 2 == 0))

    arch_row = doc_start + len(docs) + 3
    ws.merge_cells(start_row=arch_row, start_column=1, end_row=arch_row, end_column=MAX_COL)
    arc = ws.cell(row=arch_row, column=1,
        value="Archive Location:  [PLM System] -> Projects -> SPINE-CAGE-TRANSFER-2024 -> FMEA -> PFMEA-SPINE-001_RevA.xlsx")
    arc.font = Font(name="Arial", size=10, italic=True, color=NAVY)
    arc.fill = color_fill("EBF3FB")
    arc.alignment = left_align(wrap=False)
    arc.border = thin_border()

    write_footer(ws, arch_row + 2, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["C"].width = 48
    return ws

# ---------------------------------------------------------------------------
# WORKBOOK PROPERTIES
# ---------------------------------------------------------------------------

def set_workbook_properties(wb):
    wb.properties.title   = "pFMEA – PEEK Spinal Cage (AIAG-VDA 2019)"
    wb.properties.subject = "Process FMEA – Class II Medical Device"
    wb.properties.creator = "Manufacturing Transfer PE"
    wb.properties.company = "[Your Organization]"
    wb.properties.keywords = "pFMEA, AIAG-VDA, ISO 13485, 21 CFR 820, PEEK, Spinal Implant"
    wb.properties.description = (
        "AIAG-VDA 2019 pFMEA for Injection Molding of PEEK Lumbar Interbody Fusion Cage. "
        "Class II 510(k) device. Manufacturing Transfer project."
    )

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def generate_pfmea(output_path: str = "pfmea_spinal_peek_cage.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    print("Building AIAG-VDA 2019 pFMEA workbook – PEEK Spinal Cage ...")

    build_sheet1(wb)
    print("  OK  Sheet 1: Scope & Planning")

    build_sheet2(wb)
    print("  OK  Sheet 2: Structure Analysis")

    build_sheet3(wb)
    print("  OK  Sheet 3: Function Analysis")

    build_sheet4(wb)
    print("  OK  Sheet 4: Failure Analysis")

    build_sheet5(wb)
    print("  OK  Sheet 5: Risk Rating (AIAG-VDA AP)")

    build_sheet6(wb)
    print("  OK  Sheet 6: Optimization / Actions")

    build_sheet7(wb)
    print("  OK  Sheet 7: Results Summary")

    set_workbook_properties(wb)
    wb.save(output_path)
    print(f"\nDone -- Workbook saved -> {output_path}")


if __name__ == "__main__":
    import sys as _sys
    out = _sys.argv[1] if len(_sys.argv) > 1 else "pfmea_spinal_peek_cage.xlsx"
    generate_pfmea(out)
