"""
foremode.py — data-driven AIAG-VDA 2019 pFMEA generator
====================================================
One engine, many scenarios. Each scenario is a YAML file under scenarios/;
this renders any of them to a formatted Excel workbook (and, via export.py,
to Word/PDF). Action Priority and all Sheet-7 metrics are COMPUTED from the
source data — never hand-stored — so a workbook can't drift out of internal
consistency the way the old per-scenario scripts did.

Usage:
    python foremode.py list
    python foremode.py generate cnc_femoral_stem
    python foremode.py generate scenarios/cnc_femoral_stem.yaml --format all
    python foremode.py generate sterile_packaging --iso14971 -o out/pkg
    python foremode.py changelog --from old.yaml --to new.yaml

Requires: openpyxl, PyYAML  (+ python-docx, reportlab for --format docx/pdf/all)
"""

import argparse
import sys
from datetime import date
from pathlib import Path

import yaml
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SCENARIOS_DIR = Path(__file__).parent / "scenarios"

# --- palette (GMP / medical-device aesthetic) ------------------------------
NAVY, WHITE, LIGHT_GRAY = "1F3864", "FFFFFF", "F2F2F2"
RED_FILL, AMBER_FILL, GREEN_FILL = "C00000", "FFC000", "92D050"
TITLE_GRAY = "D9D9D9"
AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

# --- style helpers ---------------------------------------------------------
def hdr_font(size=11): return Font(name="Arial", bold=True, color=WHITE, size=size)
def body_font(bold=False, size=10): return Font(name="Arial", bold=bold, size=size)
def navy_fill(): return PatternFill("solid", fgColor=NAVY)
def gray_fill(): return PatternFill("solid", fgColor=LIGHT_GRAY)
def white_fill(): return PatternFill("solid", fgColor=WHITE)
def color_fill(c): return PatternFill("solid", fgColor=c)
def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left(wrap=True): return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row, columns):
    for ci, label in enumerate(columns, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font, c.fill, c.alignment, c.border = hdr_font(), navy_fill(), center(True), thin_border()

def apply_data_row(ws, row, values, alt=False):
    fill = gray_fill() if alt else white_fill()
    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font, c.fill, c.alignment, c.border = body_font(), fill, left(), thin_border()

def write_section_title(ws, row, title, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=13, color=NAVY)
    c.fill, c.alignment, c.border = color_fill(TITLE_GRAY), center(), thin_border()

def write_footer(ws, row, max_col, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", italic=True, size=9, color="595959")
    c.alignment, c.fill = center(), color_fill("EFEFEF")

def autofit(ws, min_w=12, max_w=50):
    for col in ws.columns:
        n = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(n + 4, min_w), max_w)


# --- AIAG-VDA 2019 Action Priority -----------------------------------------
def get_action_priority(s, o, d):
    """Return ('H'|'M'|'L', rationale). Detection-weighted for patient safety.

    NOTE: this is a documented, transparent simplification of the AIAG-VDA 2019
    AP logic — auditable by design, unlike black-box AI tools. It is NOT the full
    1000-cell handbook lookup table; see docs/ap-logic.md.
    """
    for name, v in (("S", s), ("O", o), ("D", d)):
        if not (isinstance(v, int) and 1 <= v <= 10):
            raise ValueError(f"{name} must be an integer 1-10, got {v!r}")
    if s >= 9:
        if d >= 7: return "H", "S>=9 + D>=7: undetected critical failure - patient risk"
        if d >= 4: return ("H", "S>=9 + D=4-6 + O>=4: frequent critical w/ weak detection") if o >= 4 \
                    else ("M", "S>=9 + D=4-6 + O<4: rare critical but detection acceptable")
        return "M", "S>=9 + D<=3: strong detection mitigates critical severity"
    if s >= 7:
        if d >= 7: return ("H", "S=7-8 + D>=7 + O>=4: high-severity, frequent, poorly detected") if o >= 4 \
                    else ("M", "S=7-8 + D>=7 + O<4: high-severity but infrequent")
        if d >= 4: return ("M", "S=7-8 + D=4-6 + O>=4: moderate risk profile") if o >= 4 \
                    else ("L", "S=7-8 + D=4-6 + O<4: controlled risk")
        return "L", "S=7-8 + D<=3: detection adequate for severity level"
    if d >= 7 and o >= 6:
        return "M", "S<=6 + high O & D: frequency/detectability concern"
    return "L", "Lower severity with adequate controls"


# --- scenario loading ------------------------------------------------------
def scenario_dirs():
    """Where to look for scenarios: ./scenarios (user's own) then the bundled set."""
    dirs, seen = [], set()
    for d in (Path.cwd() / "scenarios", SCENARIOS_DIR):
        if d not in seen:
            dirs.append(d); seen.add(d)
    return dirs

def load_scenario(ref):
    p = Path(ref)
    if not p.exists():
        for d in scenario_dirs():
            cand = d / f"{ref}.yaml"
            if cand.exists():
                p = cand; break
    if not p.exists():
        sys.exit(f"Scenario not found: {ref} (looked in {', '.join(map(str, scenario_dirs()))})")
    sc = yaml.safe_load(p.read_text(encoding="utf-8"))
    today = str(date.today())
    for row in sc.get("scope", []):
        if len(row) == 2 and row[1] == "auto":
            row[1] = today
    return sc, p


def compute_metrics(sc):
    """Derive Sheet-7 metrics from ratings + actions so they can never drift."""
    ratings = sc["ratings"]
    initial = {r[0]: get_action_priority(r[1], r[2], r[3])[0] for r in ratings}
    revised = dict(initial)
    for a in sc.get("actions", []):
        s, o, d = a["revised"]
        revised[a["mode"]] = get_action_priority(s, o, d)[0]
    def count(d, ap): return sum(1 for v in d.values() if v == ap)
    actions = sc.get("actions", [])
    return [
        ("Total Failure Modes Analyzed",        len(ratings)),
        ("Initial High (H) Action Priority",    count(initial, "H")),
        ("High AP After Actions Implemented",   count(revised, "H")),
        ("Medium (M) AP - Initial",             count(initial, "M")),
        ("Medium (M) AP - After Actions",       count(revised, "M")),
        ("Low (L) AP - After Actions",          count(revised, "L")),
        ("Total Corrective Actions Assigned",   len(actions)),
        ("Actions with Owner Assigned",         sum(1 for a in actions if a.get("owner"))),
        ("Actions with Target Date",            sum(1 for a in actions if a.get("target"))),
    ]


# --- sheet builders --------------------------------------------------------
def _kv_sheet(wb, name, title, rows, footer):
    ws = wb.create_sheet(name); ws.row_dimensions[1].height = 30
    write_section_title(ws, 1, title, 4)
    apply_header_row(ws, 2, ["Field", "Detail", "", ""])
    for i, (field, value) in enumerate(rows, start=3):
        fill = gray_fill() if i % 2 == 0 else white_fill()
        c1 = ws.cell(row=i, column=1, value=field)
        c1.font, c1.fill, c1.alignment, c1.border = body_font(True), fill, left(False), thin_border()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font, c2.fill, c2.alignment, c2.border = body_font(), fill, left(False), thin_border()
    write_footer(ws, len(rows) + 4, 4, footer)
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 60
    return ws

def _table_sheet(wb, name, title, columns, rows, footer, sev_col=None):
    ws = wb.create_sheet(name)
    write_section_title(ws, 1, title, len(columns))
    apply_header_row(ws, 2, columns)
    for i, row in enumerate(rows, start=3):
        apply_data_row(ws, i, row, alt=(i % 2 == 0))
        if sev_col is not None:
            sev = row[sev_col]
            cell = ws.cell(row=i, column=sev_col + 1)
            if isinstance(sev, int) and sev >= 9:
                cell.fill = color_fill(RED_FILL); cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            elif isinstance(sev, int) and sev >= 7:
                cell.fill = color_fill(AMBER_FILL); cell.font = body_font(True)
    write_footer(ws, len(rows) + 4, len(columns), footer)
    autofit(ws)
    return ws

def build_risk_sheet(wb, sc, footer):
    ws = wb.create_sheet("5. Risk Rating")
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value="AIAG-VDA 2019 Action Priority (AP) -- NOT Traditional RPN")
    t.font = Font(name="Arial", bold=True, size=12, color=WHITE); t.fill, t.alignment = color_fill(NAVY), center()
    ws.merge_cells("A2:G4")
    n = ws.cell(row=2, column=1, value=sc["risk_note"])
    n.font = Font(name="Arial", size=10, italic=True); n.fill = color_fill("FFF2CC")
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 70
    ws.merge_cells("A5:G5")
    sub = ws.cell(row=5, column=1,
        value="AP Lookup: S9-10+D7-10->H | S9-10+D4-6+O>=4->H | S7-8+D7-10+O>=4->H | All others->M or L")
    sub.font = Font(name="Arial", size=9, bold=True, color=NAVY); sub.fill, sub.alignment = color_fill("D9E1F2"), center()
    ws.row_dimensions[6].height = 8
    write_section_title(ws, 7, "STEP 5 - Risk Rating  |  AIAG-VDA Action Priority Table", 7)
    cols = ["Failure Mode", "S (Severity\n1-10)", "O (Occurrence\n1-10)", "D (Detection\n1-10)",
            "Action\nPriority", "AP Rationale", "Legacy RPN\n(Reference Only)"]
    apply_header_row(ws, 8, cols); ws.row_dimensions[8].height = 35
    for i, (fm, s, o, d) in enumerate((tuple(r) for r in sc["ratings"]), start=9):
        ap, rationale = get_action_priority(s, o, d)
        base = gray_fill() if i % 2 == 0 else white_fill()
        for ci, val in enumerate([fm, s, o, d, ap, rationale, s * o * d], start=1):
            c = ws.cell(row=i, column=ci, value=val)
            c.font, c.border = body_font(bold=(ci == 5)), thin_border()
            c.alignment = center(True) if ci in (2, 3, 4, 5, 7) else left()
            if ci == 5:
                c.fill = color_fill(AP_FILL[ap]); c.font = Font(name="Arial", bold=True, size=11,
                    color=WHITE if ap == "H" else "000000")
            elif ci == 7:
                c.fill = base; c.font = Font(name="Arial", size=9, italic=True, color="808080")
            else:
                c.fill = base
    write_footer(ws, len(sc["ratings"]) + 10, 7, footer); autofit(ws)
    ws.column_dimensions["F"].width = 45
    return ws

def build_optimization_sheet(wb, sc, footer):
    ws = wb.create_sheet("6. Optimization")
    write_section_title(ws, 1, "STEP 6 - Optimization / Corrective Actions  |  Risk Reduction Plan", 9)
    cols = ["Initial AP", "Failure Mode", "Action Description", "Action Owner", "Target Date",
            "Revised\nS", "Revised\nO", "Revised\nD", "Revised AP"]
    apply_header_row(ws, 2, cols); ws.row_dimensions[2].height = 35
    for i, a in enumerate(sc.get("actions", []), start=3):
        rs, ro, rd = a["revised"]
        rev_ap = get_action_priority(rs, ro, rd)[0]
        base = gray_fill() if i % 2 == 0 else white_fill()
        vals = [a["initial_ap"], a["mode"], a["action"], a.get("owner", ""), a.get("target", ""),
                rs, ro, rd, rev_ap]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=ci, value=val); c.border = thin_border()
            c.font = body_font(bold=(ci in (1, 9)))
            if ci in (1, 9):
                c.fill = color_fill(AP_FILL[val]); c.alignment = center()
                c.font = Font(name="Arial", bold=True, size=11, color=WHITE if val == "H" else "000000")
            elif ci in (6, 7, 8):
                c.fill, c.alignment = base, center()
            else:
                c.fill, c.alignment = base, left()
        ws.row_dimensions[i].height = 75
    write_footer(ws, len(sc.get("actions", [])) + 4, 9, footer)
    autofit(ws, min_w=10, max_w=60); ws.column_dimensions["C"].width = 60
    return ws

def build_results_sheet(wb, sc, footer):
    ws = wb.create_sheet("7. Results Summary")
    write_section_title(ws, 1, "STEP 7 - Results & Documentation Summary", 4)
    ws.merge_cells("A2:D2")
    mh = ws.cell(row=2, column=1, value="pFMEA Metrics Summary (auto-computed)")
    mh.font, mh.fill, mh.alignment = hdr_font(11), navy_fill(), center()
    metrics = compute_metrics(sc)
    for i, (label, value) in enumerate(metrics, start=3):
        fill = gray_fill() if i % 2 == 0 else white_fill()
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        lc = ws.cell(row=i, column=1, value=label)
        lc.font, lc.fill, lc.alignment, lc.border = body_font(True), fill, left(False), thin_border()
        vc = ws.cell(row=i, column=4, value=value)
        vc.font, vc.fill, vc.alignment, vc.border = body_font(True), fill, center(), thin_border()
    ds = len(metrics) + 4
    ws.merge_cells(start_row=ds, start_column=1, end_row=ds, end_column=4)
    dh = ws.cell(row=ds, column=1, value="Linked Quality System Documents")
    dh.font, dh.fill, dh.alignment = hdr_font(), navy_fill(), center()
    apply_header_row(ws, ds + 1, ["Document Type", "Document Number", "Title / Description", "Status"])
    docs = sc["results"]["documents"]
    for i, row in enumerate(docs, start=ds + 2):
        apply_data_row(ws, i, row, alt=(i % 2 == 0))
    ar = ds + len(docs) + 3
    ws.merge_cells(start_row=ar, start_column=1, end_row=ar, end_column=4)
    arc = ws.cell(row=ar, column=1, value="Archive Location:  " + sc["results"]["archive"])
    arc.font = Font(name="Arial", size=10, italic=True, color=NAVY)
    arc.fill, arc.alignment, arc.border = color_fill("EBF3FB"), left(False), thin_border()
    write_footer(ws, ar + 2, 4, footer); autofit(ws); ws.column_dimensions["C"].width = 48
    return ws

def build_iso14971_sheet(wb, sc, footer):
    """Optional bridge: map each pFMEA failure to ISO 14971 risk-file fields.

    The market gap nobody fills — AIAG-VDA pFMEA -> ISO 14971. Best-effort:
    pulls explicit iso14971 fields from each failure when present, else seeds
    a template row from the failure's effect/severity for the QE to complete.
    """
    ws = wb.create_sheet("8. ISO 14971 Bridge")
    ws.merge_cells("A1:G1")
    b = ws.cell(row=1, column=1, value="ISO 14971 RISK BRIDGE - best-effort mapping, QE review REQUIRED")
    b.font = Font(name="Arial", bold=True, size=11, color=WHITE); b.fill, b.alignment = color_fill(NAVY), center()
    cols = ["Failure Mode (pFMEA)", "Hazard", "Hazardous Situation", "Harm",
            "Severity (S)", "P1 (Prob. of hazardous situation)", "P2 (Prob. of harm)"]
    apply_header_row(ws, 2, cols)
    for i, f in enumerate(sc["failures"], start=3):
        iso = f.get("iso14971", {})
        row = [f["mode"], iso.get("hazard", "[derive]"),
               iso.get("hazardous_situation", "[derive]"),
               iso.get("harm", f["effect"]), f["severity"],
               iso.get("p1", "[assess]"), iso.get("p2", "[assess]")]
        apply_data_row(ws, i, row, alt=(i % 2 == 0))
    write_footer(ws, len(sc["failures"]) + 4, 7,
                 footer + "  |  ISO 14971:2019 - map P1xP2->Risk per device risk policy")
    autofit(ws)
    return ws


def build_workbook(sc, iso14971=False):
    wb = Workbook(); wb.remove(wb.active)
    footer = sc["footer"]
    pn = sc["process_name"]
    _kv_sheet(wb, "1. Scope & Planning",
              "STEP 1 - Scope & Planning  |  " + pn,
              [tuple(r) for r in sc["scope"]], footer)
    _table_sheet(wb, "2. Structure Analysis",
                 "STEP 2 - Structure Analysis  |  System Hierarchy Breakdown",
                 ["System Level", "Element Name", "Function Contribution"],
                 [tuple(r) for r in sc["structure"]], footer)
    _table_sheet(wb, "3. Function Analysis",
                 "STEP 3 - Function Analysis  |  Intended Functions & Requirements",
                 ["Process Element", "Intended Function", "Requirement Satisfied"],
                 [tuple(r) for r in sc["function"]], footer)
    _table_sheet(wb, "4. Failure Analysis",
                 "STEP 4 - Failure Analysis  |  Failure Modes, Effects & Causes",
                 ["Process Step", "Failure Mode", "Effect on Patient / Downstream Process",
                  "Severity (1-10)", "Root Cause", "Current Prevention Control", "Current Detection Control"],
                 [(f["step"], f["mode"], f["effect"], f["severity"], f["cause"],
                   f["prevention"], f["detection"]) for f in sc["failures"]],
                 footer, sev_col=3)
    build_risk_sheet(wb, sc, footer)
    build_optimization_sheet(wb, sc, footer)
    build_results_sheet(wb, sc, footer)
    if iso14971:
        build_iso14971_sheet(wb, sc, footer)
    p = sc.get("properties", {})
    wb.properties.title = p.get("title", pn)
    wb.properties.subject = p.get("subject", "Process FMEA")
    wb.properties.creator = "Quality Engineer"
    wb.properties.company = "[Your Organization]"
    wb.properties.keywords = p.get("keywords", "")
    wb.properties.description = p.get("description", "")
    return wb


# --- new-scenario template -------------------------------------------------
TEMPLATE = """\
# pFMEA scenario — edit the values, then:  python foremode.py check {name}
#                                          python foremode.py generate {name} --format all
id: {name}
process_name: "My Process - My Device"
footer: "AIAG-VDA FMEA 2019  |  21 CFR 820  |  ISO 13485"
output: foremode_{name}

properties:
  title: "pFMEA - My Device (AIAG-VDA 2019)"
  subject: "Process FMEA"
  keywords: "pFMEA, AIAG-VDA, ISO 13485"
  description: "AIAG-VDA 2019 pFMEA for <process>."

scope:                       # [Field, Detail] rows
  - ["Process Name", "My Process - My Device"]
  - ["Device Classification", "Class II - 510(k)"]
  - ["Standard Applied", "AIAG-VDA FMEA 4th Edition 2019"]
  - ["FMEA Number", "PFMEA-XXX-001"]
  - ["Revision", "Rev A"]
  - ["Date", "auto"]         # "auto" -> today's date
  - ["Status", "Draft - Internal Review"]

structure:                   # [System Level, Element, Function Contribution]
  - ["Level 1 - System", "Manufacturing Cell", "Production environment"]
  - ["Level 3 - Component", "Key Tool / Station", "What it does"]

function:                    # [Process Element, Intended Function, Requirement Satisfied]
  - ["Key Tool / Station", "Achieve spec X within tolerance Y", "Drawing / spec reference"]

failures:                    # severity is an integer 1-10
  - step: "Process Step A"
    mode: "Failure mode description"
    effect: "Effect on patient / downstream process"
    severity: 8
    cause: "Root cause"
    prevention: "Current prevention control"
    detection: "Current detection control"

risk_note: >-
  Why Action Priority (not RPN) matters for this process / device.

ratings:                     # [failure mode, S, O, D]  (all 1-10; AP is computed)
  - ["Failure mode description", 8, 3, 5]

actions:                     # corrective actions; revised is [S, O, D]
  - initial_ap: M
    mode: "Failure mode description"
    action: "What will be done to reduce risk"
    owner: "Quality Engineer"
    target: "2026-Q4"
    revised: [8, 2, 3]

results:
  documents:                 # [Type, Number, Title, Status]
    - ["Control Plan", "CP-XXX-001 Rev A", "Process Control Plan", "In Review"]
  archive: "[PLM System] -> Projects -> XXX -> FMEA -> PFMEA-XXX-001_RevA.xlsx"
"""

REQUIRED_KEYS = ["process_name", "scope", "structure", "function",
                 "failures", "risk_note", "ratings", "actions", "results"]


def validate_scenario(sc):
    """Return a list of human-readable errors ([] means valid)."""
    errs = []
    for k in REQUIRED_KEYS:
        if k not in sc:
            errs.append(f"missing required key: {k}")
    if errs:
        return errs  # structural keys missing — deeper checks would just be noise

    def sod_ok(label, s, o, d):
        for n, v in (("S", s), ("O", o), ("D", d)):
            if not (isinstance(v, int) and 1 <= v <= 10):
                errs.append(f"{label}: {n} must be int 1-10, got {v!r}")

    for i, r in enumerate(sc["ratings"]):
        if not (isinstance(r, list) and len(r) == 4):
            errs.append(f"ratings[{i}] must be [mode, S, O, D]"); continue
        sod_ok(f"ratings[{i}] '{r[0]}'", r[1], r[2], r[3])
    for i, a in enumerate(sc["actions"]):
        if a.get("initial_ap") not in ("H", "M", "L"):
            errs.append(f"actions[{i}]: initial_ap must be H/M/L")
        rev = a.get("revised")
        if not (isinstance(rev, list) and len(rev) == 3):
            errs.append(f"actions[{i}] '{a.get('mode')}': revised must be [S, O, D]")
        else:
            sod_ok(f"actions[{i}] revised", *rev)
    for i, f in enumerate(sc["failures"]):
        for k in ("step", "mode", "effect", "severity", "cause", "prevention", "detection"):
            if k not in f:
                errs.append(f"failures[{i}]: missing '{k}'")
        if "severity" in f and not (isinstance(f["severity"], int) and 1 <= f["severity"] <= 10):
            errs.append(f"failures[{i}] '{f.get('mode')}': severity must be int 1-10")
    return errs


# --- CLI -------------------------------------------------------------------
def cmd_list(_):
    seen = set()
    for d in scenario_dirs():
        for p in sorted(d.glob("*.yaml")):
            if p.stem in seen:
                continue
            seen.add(p.stem)
            sc = yaml.safe_load(p.read_text(encoding="utf-8"))
            print(f"  {p.stem:<24} {sc.get('process_name','')}")
    if not seen:
        print("  (no scenarios found — create one with: foremode new <name>)")

def cmd_new(args):
    target = Path("scenarios") / f"{args.name}.yaml"
    if target.exists() and not args.force:
        sys.exit(f"{target} already exists (use --force to overwrite)")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(TEMPLATE.format(name=args.name), encoding="utf-8")
    print(f"created {target}\n  edit it, then:  python foremode.py check {args.name}")

def cmd_check(args):
    sc, path = load_scenario(args.scenario)
    errs = validate_scenario(sc)
    fatal = [e for e in errs if not e.startswith("WARNING")]
    for e in errs:
        print(("  ! " if e.startswith("WARNING") else "  x ") + e)
    if fatal:
        sys.exit(f"{path.name}: {len(fatal)} error(s)")
    print(f"{path.name}: OK ({len(sc['ratings'])} failure modes, {len(sc['actions'])} actions)"
          + (f", {len(errs)} warning(s)" if errs else ""))

def cmd_draft(args):
    """Optional: draft a scenario YAML from free text via a local, offline LLM."""
    try:
        import llm
    except ImportError:
        sys.exit("LLM extras not available. Install with: pip install 'foremode[llm]'")
    out = Path("scenarios") / f"{args.name}.yaml"
    out.parent.mkdir(parents=True, exist_ok=True)
    llm.draft_scenario(args.describe, model=args.model, endpoint=args.endpoint,
                       scenarios_dir=SCENARIOS_DIR, use_rag=args.rag,
                       out_path=out, force=args.force)
    print(f"drafted {out} (DRAFT — review S/O/D before: python foremode.py check {args.name})")

def cmd_generate(args):
    sc, path = load_scenario(args.scenario)
    base = Path(args.output) if args.output else Path(sc.get("output", path.stem))
    base = base.with_suffix("")
    base.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(sc, iso14971=args.iso14971)
    xlsx = base.with_suffix(".xlsx"); wb.save(xlsx); print(f"wrote {xlsx}")
    if args.format in ("docx", "pdf", "all"):
        import export
        sheets = export.read_sheets(xlsx)
        if args.format in ("docx", "all"):
            out = base.with_suffix(".docx"); export.to_docx(sheets, out); print(f"wrote {out}")
        if args.format in ("pdf", "all"):
            out = base.with_suffix(".pdf"); export.to_pdf(sheets, out); print(f"wrote {out}")

def cmd_changelog(args):
    """Diff two scenario YAMLs into an auditable change summary (AIAG-VDA Step 7)."""
    a = yaml.safe_load(Path(args.from_).read_text(encoding="utf-8"))
    b = yaml.safe_load(Path(args.to).read_text(encoding="utf-8"))
    ra = {r[0]: tuple(r[1:]) for r in a["ratings"]}
    rb = {r[0]: tuple(r[1:]) for r in b["ratings"]}
    print(f"pFMEA changelog: {Path(args.from_).name} -> {Path(args.to).name}\n")
    for m in rb.keys() - ra.keys(): print(f"  + ADDED failure mode: {m} {rb[m]}")
    for m in ra.keys() - rb.keys(): print(f"  - REMOVED failure mode: {m}")
    for m in ra.keys() & rb.keys():
        if ra[m] != rb[m]:
            print(f"  ~ RE-RATED {m}: S/O/D {ra[m]} -> {rb[m]} "
                  f"(AP {get_action_priority(*ra[m])[0]} -> {get_action_priority(*rb[m])[0]})")

def main(argv=None):
    ap = argparse.ArgumentParser(description="AIAG-VDA 2019 pFMEA generator")
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("list").set_defaults(func=cmd_list)
    n = sub.add_parser("new", help="scaffold a new scenario YAML from a template")
    n.add_argument("name", help="scenario id (creates scenarios/<name>.yaml)")
    n.add_argument("--force", action="store_true", help="overwrite if it exists")
    n.set_defaults(func=cmd_new)
    ck = sub.add_parser("check", help="validate a scenario before generating")
    ck.add_argument("scenario", help="scenario id or path to a .yaml")
    ck.set_defaults(func=cmd_check)
    d = sub.add_parser("draft", help="draft a scenario from free text via a local offline LLM (optional)")
    d.add_argument("name", help="scenario id to create (scenarios/<name>.yaml)")
    d.add_argument("--describe", required=True, metavar="TEXT", help="free-text process description")
    d.add_argument("--model", default="llama3.1:8b", help="local model name")
    d.add_argument("--endpoint", default="http://localhost:11434", help="Ollama or OpenAI-compatible URL")
    d.add_argument("--rag", action="store_true", help="ground the draft on the closest existing scenario")
    d.add_argument("--force", action="store_true", help="overwrite if it exists")
    d.set_defaults(func=cmd_draft)
    g = sub.add_parser("generate")
    g.add_argument("scenario", help="scenario id (in scenarios/) or path to a .yaml")
    g.add_argument("--format", choices=["xlsx", "docx", "pdf", "all"], default="xlsx")
    g.add_argument("--iso14971", action="store_true", help="add ISO 14971 risk-bridge sheet")
    g.add_argument("-o", "--output", help="output basename (extension ignored)")
    g.set_defaults(func=cmd_generate)
    c = sub.add_parser("changelog")
    c.add_argument("--from", dest="from_", required=True)
    c.add_argument("--to", required=True)
    c.set_defaults(func=cmd_changelog)
    args = ap.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
