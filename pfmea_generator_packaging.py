"""
pfmea_generator_packaging.py
============================
Generates a complete AIAG-VDA 2019 pFMEA Excel workbook for a
Sterile Barrier Packaging Seal Integrity process — Tyvek-Film Pouch
Heat Sealing for Class III implants.

Portfolio piece: Manufacturing Transfer / Quality Engineering showcase.
Regulatory context: 21 CFR 820, ISO 13485:2016, ISO 11607-1:2019.

Author : [Your Name]
Version: 1.0.0
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# PALETTE
# ---------------------------------------------------------------------------
NAVY       = "1F3864"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_FILL   = "FF0000"
AMBER_FILL = "FFC000"
GREEN_FILL = "92D050"
TITLE_GRAY = "D9D9D9"

FOOTER_TEXT  = "AIAG-VDA FMEA 2019  |  21 CFR 820  |  ISO 13485  |  ISO 11607"
PROCESS_NAME = "Sterile Barrier Packaging Seal – Tyvek-Film Pouch (Class III Implants)"

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def hdr_font(size=11):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size)

def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def color_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, columns):
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = hdr_font()
        cell.fill = navy_fill()
        cell.alignment = center_align(wrap=True)
        cell.border = thin_border()

def apply_data_row(ws, row_num, values, alt=False):
    fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = body_font()
        cell.fill = fill
        cell.alignment = left_align()
        cell.border = thin_border()

def write_section_title(ws, row_num, title, max_col):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=13, color=NAVY)
    cell.fill = color_fill(TITLE_GRAY)
    cell.alignment = center_align()
    cell.border = thin_border()

def write_footer(ws, row_num, max_col):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=max_col)
    cell = ws.cell(row=row_num, column=1, value=FOOTER_TEXT)
    cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    cell.alignment = center_align()
    cell.fill = color_fill("EFEFEF")

def autofit_columns(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

# ---------------------------------------------------------------------------
# AIAG-VDA 2019 ACTION PRIORITY
# ---------------------------------------------------------------------------

def get_action_priority(s, o, d):
    if s >= 9:
        if d >= 7:
            return "H", "S>=9 + D>=7: undetected critical failure – patient risk"
        elif 4 <= d <= 6:
            if o >= 4:
                return "H", "S>=9 + D=4-6 + O>=4: frequent critical w/ weak detection"
            else:
                return "M", "S>=9 + D=4-6 + O<4: rare critical but detection acceptable"
        else:
            return "M", "S>=9 + D<=3: strong detection mitigates critical severity"
    elif s >= 7:
        if d >= 7:
            if o >= 4:
                return "H", "S=7-8 + D>=7 + O>=4: high-severity, frequent, poorly detected"
            else:
                return "M", "S=7-8 + D>=7 + O<4: high-severity but infrequent"
        elif 4 <= d <= 6:
            if o >= 4:
                return "M", "S=7-8 + D=4-6 + O>=4: moderate risk profile"
            else:
                return "L", "S=7-8 + D=4-6 + O<4: controlled risk"
        else:
            return "L", "S=7-8 + D<=3: detection adequate for severity level"
    else:
        if d >= 7 and o >= 6:
            return "M", "S<=6 + high O & D: frequency/detectability concern"
        else:
            return "L", "Lower severity with adequate controls"

# ---------------------------------------------------------------------------
# SHEET 1 – SCOPE & PLANNING
# ---------------------------------------------------------------------------

def build_sheet1(wb):
    ws = wb.create_sheet("1. Scope & Planning")
    MAX_COL = 4
    ws.row_dimensions[1].height = 30

    write_section_title(ws, 1, "STEP 1 – Scope & Planning  |  " + PROCESS_NAME, MAX_COL)

    fields = [
        ("Process Name",          PROCESS_NAME),
        ("Scope Start",           "Pouching / Loading of Sterilized Implant"),
        ("Scope End",             "Sealing, Inspection, Labeling & Distribution Release"),
        ("Device Classification", "Class III – PMA support process (21 CFR 814)"),
        ("Standard Applied",      "AIAG-VDA FMEA 4th Edition 2019"),
        ("Regulatory Framework",  "21 CFR 820 (QSR)  |  ISO 13485:2016  |  ISO 11607-1:2019"),
        ("Device Description",    "Sterile Tyvek-Film Pouch – Primary Sterile Barrier for Class III Implants"),
        ("Team Members",          "Packaging Engineer, Quality Engineer, Sterilization Specialist, Supplier Quality"),
        ("FMEA Facilitator",      "[Your Name] – Manufacturing Transfer Project Engineer"),
        ("FMEA Number",           "PFMEA-PKG-001"),
        ("Revision",              "Rev A"),
        ("Date",                  str(date.today())),
        ("Next Review Date",      "Annual review per SOP-QE-022 or upon process change"),
        ("Status",                "Draft – Internal Review"),
    ]

    apply_header_row(ws, 2, ["Field", "Detail", "", ""])

    for i, (field, value) in enumerate(fields, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        c1 = ws.cell(row=i, column=1, value=field)
        c1.font = body_font(bold=True)
        c1.fill = fill
        c1.alignment = left_align(wrap=False)
        c1.border = thin_border()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=MAX_COL)
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font = body_font()
        c2.fill = fill
        c2.alignment = left_align(wrap=False)
        c2.border = thin_border()

    write_footer(ws, len(fields) + 4, MAX_COL)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    return ws

# ---------------------------------------------------------------------------
# SHEET 2 – STRUCTURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet2(wb):
    ws = wb.create_sheet("2. Structure Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 2 – Structure Analysis  |  System Hierarchy Breakdown", MAX_COL)
    apply_header_row(ws, 2, ["System Level", "Element Name", "Function Contribution"])

    rows = [
        ("Level 1 – System",     "Sterile Packaging Line",            "End-to-end sterile barrier formation and validation"),
        ("Level 2 – Subsystem",  "Rotary Heat Sealer",                "Forms hermetic seal on three open edges of Tyvek-film pouch"),
        ("Level 3 – Component",  "Sealing Jaw / Platen",              "Applies controlled heat and pressure to create seal bond"),
        ("Level 3 – Component",  "Temperature Controller (PID)",      "Maintains seal jaw within validated temperature window"),
        ("Level 3 – Component",  "Dwell Timer",                       "Controls contact time for complete seal formation"),
        ("Level 3 – Component",  "Pouch Stock (Tyvek + Film)",        "Provides sterile barrier material with validated bond surface"),
        ("Level 3 – Component",  "Operator / Packaging Technician",   "Loads implant, positions pouch, initiates seal cycle"),
        ("Level 4 – Interface",  "Seal Integrity Test Station",       "Bubble emission (ASTM F2096) and peel strength verification"),
        ("Level 4 – Interface",  "Label Applicator / eDHR",           "Applies UDI label; records lot, operator, and seal params"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 3 – FUNCTION ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet3(wb):
    ws = wb.create_sheet("3. Function Analysis")
    MAX_COL = 3

    write_section_title(ws, 1, "STEP 3 – Function Analysis  |  Intended Functions & Requirements", MAX_COL)
    apply_header_row(ws, 2, ["Process Element", "Intended Function", "Requirement Satisfied"])

    rows = [
        ("Rotary Heat Sealer",
         "Form a continuous, channel-free hermetic seal on three sides; seal width >= 6 mm per ISO 11607-1",
         "ISO 11607-1 Annex A: seal width and channel-defect acceptance criteria"),
        ("Temperature Controller (PID)",
         "Maintain sealing jaw at 175 +/-5 C (validated window per OQ); alarm on deviation > 3 C",
         "Seal OQ temperature mapping; process validation PQ-PKG-SEAL-001"),
        ("Dwell Timer",
         "Hold seal pressure for 1.2 +/-0.1 s (validated); prevent under-dwell causing weak seal",
         "Seal strength OQ data; minimum peel force >= 1.5 N/15 mm per ASTM F88"),
        ("Pouch Stock (Tyvek + Film)",
         "Maintain microbial barrier and seal-ability; stored per supplier IFU (humidity/temp controlled)",
         "ISO 11607-1 material qualification; supplier CoA per receiving SOP"),
        ("Operator / Packaging Tech",
         "Load single implant per pouch, orient correctly, position seal edge within fixture guide",
         "Work instruction WI-PKG-001; operator qualification and annual re-certification"),
        ("Seal Integrity Test Station",
         "Detect seal channels >1 mm via bubble emission (ASTM F2096); measure peel strength per ASTM F88",
         "ISO 11607-1 Annex A2.2; AQL sampling plan SP-PKG-02"),
        ("Label Applicator / eDHR",
         "Apply correct UDI label with lot/SN; capture seal temp, dwell, operator ID in DHR",
         "21 CFR 830 UDI; 21 CFR 820.184 DHR; ISO 13485 section 7.5.3 traceability"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 4 – FAILURE ANALYSIS
# ---------------------------------------------------------------------------

def build_sheet4(wb):
    ws = wb.create_sheet("4. Failure Analysis")
    MAX_COL = 7

    write_section_title(ws, 1, "STEP 4 – Failure Analysis  |  Failure Modes, Effects & Causes", MAX_COL)

    columns = [
        "Process Step", "Failure Mode",
        "Effect on Patient / Downstream Process",
        "Severity (1-10)", "Root Cause",
        "Current Prevention Control", "Current Detection Control"
    ]
    apply_header_row(ws, 2, columns)

    rows = [
        ("Heat Sealing",
         "Incomplete / channel seal defect (> 1 mm channel)",
         "Loss of sterile barrier; microbial ingress; patient infection from non-sterile implant",
         9,
         "Jaw misalignment; pouch wrinkle in seal zone; temperature excursion; worn platen",
         "Seal jaw alignment verification at setup; temperature setpoint validated per OQ",
         "Visual inspection per WI-VIS-004 – channels < 1 mm not reliably detected visually (GAP)"),
        ("Seal Integrity",
         "Pinhole or micro-leak in film layer (microbial ingress path)",
         "Sterility breach; patient infection; implant recall and field safety corrective action",
         10,
         "Film puncture from implant edge; handling damage; raw material defect",
         "Pouch edge guard; gentle handling SOP; incoming film inspection",
         "Bubble emission test samples only (ASTM F2096) – not 100% per lot (DETECTION GAP)"),
        ("Temperature Control",
         "Seal temperature out of validated range (< 170 C or > 180 C)",
         "Under-seal: weak bond, field delamination risk. Over-seal: Tyvek fiber damage, barrier compromise",
         8,
         "PID controller drift; thermocouple degradation; power supply fluctuation",
         "PID calibration per PM schedule; operator pre-shift temperature verification",
         "Process monitoring via controller display – no independent thermocouple alarm (GAP)"),
        ("Dwell Time",
         "Incorrect dwell time (operator override or timer fault)",
         "Under-dwell: weak seal, peel strength below 1.5 N/15 mm; sterile barrier at risk in distribution",
         8,
         "Operator bypasses timer; controller fault; recipe not locked",
         "Timer setpoint in validated recipe; operator training",
         "Peel strength sampled 5 pouches/lot per ASTM F88 – sub-lot variation not captured"),
        ("Labeling",
         "UDI label / content mismatch (wrong lot or implant size on label)",
         "Incorrect device implanted; patient harm; regulatory non-conformance (21 CFR 830)",
         8,
         "Manual label selection; look-alike label stock; no barcode verification at point of use",
         "Label control SOP; label segregation; 2-person verification",
         "Final QA label review before release – human verification only (GAP)"),
        ("Post-Seal Handling",
         "Package breach / seal peel during EtO sterilization or distribution transit",
         "Sterility loss at point of use; patient infection; product recall",
         9,
         "Over-stacking in sterilization load; inadequate secondary packaging cushioning",
         "Stacking height limit in sterilization SOP; secondary packaging spec",
         "Post-EtO visual inspection (sampling only); distribution simulation testing per ASTM D4169 annual"),
    ]

    for i, row_data in enumerate(rows, start=3):
        apply_data_row(ws, i, row_data, alt=(i % 2 == 0))
        s_cell = ws.cell(row=i, column=4)
        if isinstance(row_data[3], int) and row_data[3] >= 9:
            s_cell.fill = color_fill(RED_FILL)
            s_cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        elif isinstance(row_data[3], int) and row_data[3] >= 7:
            s_cell.fill = color_fill(AMBER_FILL)
            s_cell.font = body_font(bold=True)

    write_footer(ws, len(rows) + 4, MAX_COL)
    autofit_columns(ws)
    return ws

# ---------------------------------------------------------------------------
# SHEET 5 – RISK RATING
# ---------------------------------------------------------------------------

def build_sheet5(wb):
    ws = wb.create_sheet("5. Risk Rating")
    MAX_COL = 7

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
        value="  AIAG-VDA 2019 Action Priority (AP) -- NOT Traditional RPN")
    t.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    t.fill = color_fill(NAVY)
    t.alignment = center_align()

    note = (
        "AIAG-VDA 2019 replaces RPN (S x O x D) with Action Priority (AP: High / Medium / Low). "
        "For sterile packaging, the consequence of an undetected seal defect is a non-sterile implant "
        "reaching the patient — a catastrophic outcome regardless of defect frequency. "
        "ISO 11607 requires validated sterile barrier integrity; AP prioritizes detection gaps accordingly. "
        "Any Severity >= 9 failure with Detection >= 7 is automatically High priority — "
        "RPN arithmetic can mask this by averaging a low Occurrence score."
    )
    ws.merge_cells("A2:G4")
    n = ws.cell(row=2, column=1, value=note)
    n.font = Font(name="Arial", size=10, italic=True)
    n.fill = color_fill("FFF2CC")
    n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 72

    ws.merge_cells("A5:G5")
    sub = ws.cell(row=5, column=1,
        value="AP Lookup: S9-10+D7-10->H | S9-10+D4-6+O>=4->H | S7-8+D7-10+O>=4->H | All others->M or L")
    sub.font = Font(name="Arial", size=9, bold=True, color=NAVY)
    sub.fill = color_fill("D9E1F2")
    sub.alignment = center_align()
    ws.row_dimensions[6].height = 8

    write_section_title(ws, 7, "STEP 5 – Risk Rating  |  AIAG-VDA Action Priority Table", MAX_COL)

    columns = [
        "Failure Mode",
        "S  (Severity\n1-10)",
        "O  (Occurrence\n1-10)",
        "D  (Detection\n1-10)",
        "Action\nPriority",
        "AP Rationale",
        "Legacy RPN\n(Reference Only)"
    ]
    apply_header_row(ws, 8, columns)
    ws.row_dimensions[8].height = 35

    rated_failures = [
        ("Incomplete / channel seal defect",          9, 4, 6),
        ("Pinhole / micro-leak in film layer",        10, 2, 8),
        ("Seal temperature out of range",             8, 4, 5),
        ("Incorrect dwell time",                      8, 3, 3),
        ("UDI label / content mismatch",              8, 2, 4),
        ("Package breach during EtO / transit",       9, 3, 5),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (fm, s, o, d) in enumerate(rated_failures, start=9):
        ap, rationale = get_action_priority(s, o, d)
        rpn = s * o * d
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        row_vals = [fm, s, o, d, ap, rationale, rpn]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font = body_font(bold=(col_idx == 5))
            cell.border = thin_border()
            cell.alignment = center_align(wrap=True) if col_idx in (2, 3, 4, 5, 7) else left_align()

            if col_idx == 5:
                cell.fill = color_fill(AP_FILL[ap])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap == "H" else "000000")
            elif col_idx == 7:
                cell.fill = base_fill
                cell.font = Font(name="Arial", size=9, italic=True, color="808080")
            else:
                cell.fill = base_fill

    write_footer(ws, len(rated_failures) + 10, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["F"].width = 45
    return ws

# ---------------------------------------------------------------------------
# SHEET 6 – OPTIMIZATION
# ---------------------------------------------------------------------------

def build_sheet6(wb):
    ws = wb.create_sheet("6. Optimization")
    MAX_COL = 9

    write_section_title(ws, 1, "STEP 6 – Optimization / Corrective Actions  |  Risk Reduction Plan", MAX_COL)

    columns = [
        "Initial AP", "Failure Mode",
        "Action Description", "Action Owner",
        "Target Date",
        "Revised\nS", "Revised\nO", "Revised\nD",
        "Revised AP"
    ]
    apply_header_row(ws, 2, columns)
    ws.row_dimensions[2].height = 35

    actions = [
        ("H",  "Incomplete / channel seal defect",
         "Implement 100% bubble emission test per ASTM F2096 (replacing sampling). "
         "Add automated peel strength tester with SPC control chart – alert at < 1.8 N/15 mm. "
         "Update Control Plan CP-PKG-001 and sampling plan SP-PKG-02. "
         "Validate test equipment per IQ/OQ IQ-SIT-001.",
         "Packaging Engineer + QE", "2026-Q3",  9, 2, 3),

        ("H",  "Pinhole / micro-leak in film layer",
         "Add dye penetration test per ISO 11607-1 Annex A2.2 to every batch AQL (n=32, AQL 0.65). "
         "Implement incoming film inspection with pin-hole detector per incoming SOP. "
         "Add implant edge guard (foam insert) to pouching fixture.",
         "Quality Engineer", "2026-Q3",  10, 2, 4),

        ("M",  "Seal temperature out of range",
         "Install independent thermocouple with dual-channel alarm: feed-hold if jaw temp deviates "
         "> 3 C from setpoint. Validate alarm response per OQ protocol OQ-SEAL-002. "
         "Reduce PID calibration interval from 12 to 6 months.",
         "Process Engineer", "2026-Q4",  8, 2, 3),

        ("M",  "Package breach during EtO / transit",
         "Add post-EtO visual inspection of all pouches before labeling (100%, not sampling). "
         "Implement ASTM D4169 distribution simulation testing semi-annually (current: annual). "
         "Add foam separator between pouch layers in sterilization cassette.",
         "Packaging + Sterilization PE", "2026-Q4",  9, 2, 3),
    ]

    AP_FILL = {"H": RED_FILL, "M": AMBER_FILL, "L": GREEN_FILL}

    for i, (init_ap, fm, action, owner, target, rs, ro, rd) in enumerate(actions, start=3):
        rev_ap, _ = get_action_priority(rs, ro, rd)
        alt = (i % 2 == 0)
        base_fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)

        for col_idx, val in enumerate(
            [init_ap, fm, action, owner, target, rs, ro, rd, rev_ap], start=1
        ):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin_border()
            cell.font = body_font(bold=(col_idx in (1, 9)))

            if col_idx in (1, 9):
                ap_val = val
                cell.fill = color_fill(AP_FILL[ap_val])
                cell.font = Font(name="Arial", bold=True, size=11,
                                 color=WHITE if ap_val == "H" else "000000")
                cell.alignment = center_align()
            elif col_idx in (6, 7, 8):
                cell.fill = base_fill
                cell.alignment = center_align()
            else:
                cell.fill = base_fill
                cell.alignment = left_align()

        ws.row_dimensions[i].height = 80

    write_footer(ws, len(actions) + 4, MAX_COL)
    autofit_columns(ws, min_width=10, max_width=60)
    ws.column_dimensions["C"].width = 60
    return ws

# ---------------------------------------------------------------------------
# SHEET 7 – RESULTS SUMMARY
# ---------------------------------------------------------------------------

def build_sheet7(wb):
    ws = wb.create_sheet("7. Results Summary")
    MAX_COL = 4

    write_section_title(ws, 1, "STEP 7 – Results & Documentation Summary", MAX_COL)

    ws.merge_cells("A2:D2")
    mhdr = ws.cell(row=2, column=1, value="pFMEA Metrics Summary")
    mhdr.font = hdr_font(size=11)
    mhdr.fill = navy_fill()
    mhdr.alignment = center_align()

    metrics = [
        ("Total Failure Modes Analyzed",         6),
        ("Initial High (H) Action Priority",     2),
        ("High AP After Actions Implemented",    0),
        ("Medium (M) AP – Initial",              4),
        ("Medium (M) AP – After Actions",        4),
        ("Low (L) AP – After Actions",           2),
        ("Total Corrective Actions Assigned",    4),
        ("Actions with Owner Assigned",          4),
        ("Actions with Target Date",             4),
    ]

    for i, (label, value) in enumerate(metrics, start=3):
        alt = (i % 2 == 0)
        fill = gray_fill() if alt else PatternFill("solid", fgColor=WHITE)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = body_font(bold=True)
        lc.fill = fill
        lc.alignment = left_align(wrap=False)
        lc.border = thin_border()
        vc = ws.cell(row=i, column=4, value=value)
        vc.font = body_font(bold=True)
        vc.fill = fill
        vc.alignment = center_align()
        vc.border = thin_border()

    doc_start = len(metrics) + 4
    ws.merge_cells(start_row=doc_start, start_column=1,
                   end_row=doc_start, end_column=MAX_COL)
    dhdr = ws.cell(row=doc_start, column=1, value="Linked Quality System Documents")
    dhdr.font = hdr_font()
    dhdr.fill = navy_fill()
    dhdr.alignment = center_align()

    apply_header_row(ws, doc_start + 1,
                     ["Document Type", "Document Number", "Title / Description", "Status"])

    docs = [
        ("Control Plan",     "CP-PKG-001 Rev A",    "Sterile Packaging – Heat Seal Control Plan",                   "In Review"),
        ("Work Instruction", "WI-PKG-001 Rev C",    "Tyvek-Film Pouch Sealing – Setup & Operation",                "Released"),
        ("Inspection Plan",  "SP-PKG-02 Rev B",     "Seal Integrity Sampling Plan – Bubble Emission & Peel Strength","Released"),
        ("OQ Protocol",      "OQ-SEAL-002",         "Heat Sealer Temperature & Dwell Time – Operational Qualification","In Progress"),
        ("PQ Protocol",      "PQ-PKG-SEAL-001",     "Process Validation – Sterile Packaging Transfer",              "Planned"),
        ("Risk Management",  "RM-PKG-2024-01",      "Packaging Risk Management File (ISO 14971)",                   "In Review"),
        ("Sterilization PQ", "PQ-ETO-HIP-001",      "EtO Sterilization Validation – Class III Implant Family",      "Released"),
        ("SOP",              "SOP-QE-022",          "pFMEA Creation and Maintenance Procedure",                     "Released"),
    ]

    for i, doc_row in enumerate(docs, start=doc_start + 2):
        apply_data_row(ws, i, doc_row, alt=(i % 2 == 0))

    arch_row = doc_start + len(docs) + 3
    ws.merge_cells(start_row=arch_row, start_column=1, end_row=arch_row, end_column=MAX_COL)
    arc = ws.cell(row=arch_row, column=1,
        value="Archive Location:  [PLM System] -> Projects -> PKG-SEAL-TRANSFER-2024 -> FMEA -> PFMEA-PKG-001_RevA.xlsx")
    arc.font = Font(name="Arial", size=10, italic=True, color=NAVY)
    arc.fill = color_fill("EBF3FB")
    arc.alignment = left_align(wrap=False)
    arc.border = thin_border()

    write_footer(ws, arch_row + 2, MAX_COL)
    autofit_columns(ws)
    ws.column_dimensions["C"].width = 48
    return ws

# ---------------------------------------------------------------------------
# WORKBOOK PROPERTIES
# ---------------------------------------------------------------------------

def set_workbook_properties(wb):
    wb.properties.title   = "pFMEA – Sterile Barrier Packaging (AIAG-VDA 2019)"
    wb.properties.subject = "Process FMEA – Sterile Packaging / ISO 11607"
    wb.properties.creator = "Manufacturing Transfer PE"
    wb.properties.company = "[Your Organization]"
    wb.properties.keywords = "pFMEA, AIAG-VDA, ISO 13485, 21 CFR 820, ISO 11607, Sterile Packaging"
    wb.properties.description = (
        "AIAG-VDA 2019 pFMEA for Tyvek-Film Pouch Heat Sealing. "
        "Supports Class III PMA implants. ISO 11607-1:2019 sterile barrier process."
    )

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def generate_pfmea(output_path: str = "pfmea_sterile_packaging.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    print("Building AIAG-VDA 2019 pFMEA workbook – Sterile Barrier Packaging ...")

    build_sheet1(wb)
    print("  OK  Sheet 1: Scope & Planning")

    build_sheet2(wb)
    print("  OK  Sheet 2: Structure Analysis")

    build_sheet3(wb)
    print("  OK  Sheet 3: Function Analysis")

    build_sheet4(wb)
    print("  OK  Sheet 4: Failure Analysis")

    build_sheet5(wb)
    print("  OK  Sheet 5: Risk Rating (AIAG-VDA AP)")

    build_sheet6(wb)
    print("  OK  Sheet 6: Optimization / Actions")

    build_sheet7(wb)
    print("  OK  Sheet 7: Results Summary")

    set_workbook_properties(wb)
    wb.save(output_path)
    print(f"\nDone -- Workbook saved -> {output_path}")


if __name__ == "__main__":
    import sys as _sys
    out = _sys.argv[1] if len(_sys.argv) > 1 else "pfmea_sterile_packaging.xlsx"
    generate_pfmea(out)
